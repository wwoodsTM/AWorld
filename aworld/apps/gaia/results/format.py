import json
import os

from openpyxl import load_workbook

# 定义文件路径
json_file_path = "./result_#4_0_165.json"
excel_file_path = "./target.xlsx"

# 确保Excel文件存在
if not os.path.exists(excel_file_path):
    print(f"Excel doesn't exists : {excel_file_path}")
    exit(1)

# 读取JSON文件
try:
    with open(json_file_path, "r", encoding="utf-8") as f:
        results = json.load(f)
except Exception as e:
    print(f"Reading JSON failed: {str(e)}")
    exit(1)

task_id_to_score = {}
for item in results:
    score_value = 1 if item.get("score") is True else 0
    task_id_to_score[item.get("task_id")] = score_value

# 加载Excel工作簿
try:
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    updated_count = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        task_id_cell = row[0]
        score_cell = row[1]

        if task_id_cell.value in task_id_to_score:
            score_cell.value = task_id_to_score[task_id_cell.value]
            updated_count += 1
        else:
            score_cell.value = -1
            updated_count += 1

    workbook.save(excel_file_path)

except Exception as e:
    print(str(e))
    exit(1)

print("Done!")
